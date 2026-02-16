"""
Database Builder - Creates Multi-Sheet Excel Database

Builds a comprehensive database with 4 sheets:
1. Master Database - All programs
2. Active Programs - Filtered (ACTIVE, FEDERAL)
3. Cleanup Required - DUPLICATE, NON_INCENTIVE, MISSING_LINK, EXPIRED, HALLUCINATION
4. Executive Summary - Statistics, key findings, next steps
"""

import os
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class DatabaseStats:
    """Statistics about the database"""
    total_programs: int
    active_programs: int
    federal_programs: int
    duplicate_programs: int
    expired_programs: int
    non_incentive_programs: int
    missing_link_programs: int
    hallucination_programs: int
    coverage_score: float
    populations_covered: List[str]
    program_types_found: List[str]


class DatabaseBuilder:
    """
    Builds multi-sheet Excel database from discovered programs.
    """

    # Status tags and their meanings
    STATUS_DEFINITIONS = {
        "ACTIVE": "Verified active program with direct employer benefit",
        "FEDERAL": "Federal program with state implementation",
        "DUPLICATE": "Duplicate of another program in database",
        "EXPIRED": "Program no longer active",
        "NON-INCENTIVE": "No direct employer benefit",
        "MISSING-LINK": "Cannot verify - missing source URL",
        "HALLUCINATION": "Likely fabricated - remove from database",
        "REVIEW": "Needs manual review"
    }

    # Columns for each sheet
    MASTER_COLUMNS = [
        "Program_ID", "Program_Name", "Agency", "Status_Tag", "Benefit_Type",
        "Jurisdiction", "Max_Value", "Target_Populations", "Description",
        "Official_Source_URL", "Confidence", "Classification_Reasoning", "Notes"
    ]

    ACTIVE_COLUMNS = [
        "Program_ID", "Program_Name", "Agency", "Benefit_Type", "Jurisdiction",
        "Max_Value", "Target_Populations", "Description", "Official_Source_URL"
    ]

    CLEANUP_COLUMNS = [
        "Program_ID", "Program_Name", "Status_Tag", "Issue", "Suggested_Action", "Notes"
    ]

    def __init__(self):
        self.programs: List[Dict[str, Any]] = []
        self.gaps: Optional[Dict[str, Any]] = None
        self.mental_model: Optional[Dict[str, Any]] = None

    def build_database(self, programs: List[Dict[str, Any]],
                       gaps: Optional[Dict[str, Any]] = None,
                       mental_model: Optional[Dict[str, Any]] = None,
                       output_path: str = "outputs/database.xlsx") -> str:
        """
        Build the complete database.

        Args:
            programs: List of classified programs
            gaps: Gap analysis results
            mental_model: Mental model from landscape mapping
            output_path: Path for output Excel file

        Returns:
            Path to created file
        """
        self.programs = programs
        self.gaps = gaps
        self.mental_model = mental_model

        # Assign program IDs if not present
        self._assign_program_ids()

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create sheets
        self._create_master_sheet(wb)
        self._create_active_sheet(wb)
        self._create_cleanup_sheet(wb)
        self._create_summary_sheet(wb)

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        # Save
        wb.save(output_path)
        print(f"  Database saved to: {output_path}")

        return output_path

    def _assign_program_ids(self):
        """Assign unique IDs to programs"""
        # Get state code from first program or default to XX
        state_code = "XX"
        if self.programs:
            state = self.programs[0].get("state", "")
            if state:
                # Use first two letters of state
                state_code = state[:2].upper()

        for i, program in enumerate(self.programs, 1):
            if not program.get("Program_ID"):
                program["Program_ID"] = f"{state_code}-{i:03d}"

    def _create_master_sheet(self, wb: Workbook):
        """Create Master Database sheet with all programs"""
        ws = wb.create_sheet("Master Database")

        # Prepare data
        data = []
        for prog in self.programs:
            row = {
                "Program_ID": prog.get("Program_ID", ""),
                "Program_Name": prog.get("program_name", prog.get("Program_Name", "")),
                "Agency": prog.get("agency", prog.get("Agency", "")),
                "Status_Tag": prog.get("status_tag", prog.get("Status_Tag", "")),
                "Benefit_Type": prog.get("benefit_type", ""),
                "Jurisdiction": prog.get("jurisdiction", ""),
                "Max_Value": prog.get("max_value", prog.get("verified_max_value", prog.get("Verified_Max_Value", ""))),
                "Target_Populations": self._format_populations(prog.get("target_populations", [])),
                "Description": prog.get("description", ""),
                "Official_Source_URL": prog.get("source_url", prog.get("official_source_url", prog.get("Official_Source_URL", ""))),
                "Confidence": prog.get("confidence", ""),
                "Classification_Reasoning": prog.get("classification_reasoning", ""),
                "Notes": prog.get("notes", prog.get("Notes", ""))
            }
            data.append(row)

        df = pd.DataFrame(data, columns=self.MASTER_COLUMNS)

        # Write to sheet
        self._write_dataframe_to_sheet(ws, df)

        # Apply styling
        self._style_header(ws)
        self._apply_status_colors(ws, status_col=4)  # Status_Tag is column D (4)

    def _create_active_sheet(self, wb: Workbook):
        """Create Active Programs sheet with only usable programs"""
        ws = wb.create_sheet("Active Programs")

        # Filter for ACTIVE and FEDERAL programs
        active_programs = [
            p for p in self.programs
            if p.get("status_tag", p.get("Status_Tag", "")) in ["ACTIVE", "FEDERAL"]
        ]

        # Prepare data
        data = []
        for prog in active_programs:
            row = {
                "Program_ID": prog.get("Program_ID", ""),
                "Program_Name": prog.get("program_name", prog.get("Program_Name", "")),
                "Agency": prog.get("agency", prog.get("Agency", "")),
                "Benefit_Type": prog.get("benefit_type", ""),
                "Jurisdiction": prog.get("jurisdiction", ""),
                "Max_Value": prog.get("max_value", prog.get("verified_max_value", prog.get("Verified_Max_Value", ""))),
                "Target_Populations": self._format_populations(prog.get("target_populations", [])),
                "Description": prog.get("description", ""),
                "Official_Source_URL": prog.get("source_url", prog.get("official_source_url", prog.get("Official_Source_URL", "")))
            }
            data.append(row)

        df = pd.DataFrame(data, columns=self.ACTIVE_COLUMNS)

        self._write_dataframe_to_sheet(ws, df)
        self._style_header(ws)

    def _create_cleanup_sheet(self, wb: Workbook):
        """Create Cleanup Required sheet with programs needing attention"""
        ws = wb.create_sheet("Cleanup Required")

        # Filter for programs needing cleanup
        cleanup_statuses = ["DUPLICATE", "NON-INCENTIVE", "MISSING-LINK", "EXPIRED", "HALLUCINATION", "REVIEW"]
        cleanup_programs = [
            p for p in self.programs
            if p.get("status_tag", p.get("Status_Tag", "")) in cleanup_statuses
        ]

        # Prepare data
        data = []
        for prog in cleanup_programs:
            status = prog.get("status_tag", prog.get("Status_Tag", ""))
            issue, action = self._get_cleanup_guidance(status, prog)

            row = {
                "Program_ID": prog.get("Program_ID", ""),
                "Program_Name": prog.get("program_name", prog.get("Program_Name", "")),
                "Status_Tag": status,
                "Issue": issue,
                "Suggested_Action": action,
                "Notes": prog.get("notes", prog.get("Notes", ""))
            }
            data.append(row)

        df = pd.DataFrame(data, columns=self.CLEANUP_COLUMNS)

        self._write_dataframe_to_sheet(ws, df)
        self._style_header(ws)
        self._apply_status_colors(ws, status_col=3)  # Status_Tag is column C (3)

    def _create_summary_sheet(self, wb: Workbook):
        """Create Executive Summary sheet"""
        ws = wb.create_sheet("Executive Summary")

        # Calculate stats
        stats = self._calculate_stats()

        # Header
        ws['A1'] = "EXECUTIVE SUMMARY"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

        # State info
        if self.mental_model:
            ws['A4'] = f"State: {self.mental_model.get('state', 'Unknown')}"
            ws['A4'].font = Font(bold=True)

        # Overall Stats section
        row = 6
        ws[f'A{row}'] = "PROGRAM STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        stats_data = [
            ("Total Programs", stats.total_programs),
            ("Active Programs (usable)", stats.active_programs),
            ("Federal Programs", stats.federal_programs),
            ("Duplicates", stats.duplicate_programs),
            ("Expired", stats.expired_programs),
            ("Non-Incentives", stats.non_incentive_programs),
            ("Missing Links", stats.missing_link_programs),
            ("Potential Hallucinations", stats.hallucination_programs),
        ]

        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Coverage Score
        row += 1
        ws[f'A{row}'] = "COVERAGE SCORE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "Score"
        ws[f'B{row}'] = f"{stats.coverage_score:.1f}%"
        score_color = "00FF00" if stats.coverage_score >= 80 else "FFFF00" if stats.coverage_score >= 60 else "FF0000"
        ws[f'B{row}'].fill = PatternFill(start_color=score_color, end_color=score_color, fill_type="solid")
        row += 1

        # Populations covered
        row += 1
        ws[f'A{row}'] = "POPULATIONS COVERED"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for pop in stats.populations_covered:
            ws[f'A{row}'] = f"  - {pop}"
            row += 1

        # Program types found
        row += 1
        ws[f'A{row}'] = "PROGRAM TYPES FOUND"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for ptype in stats.program_types_found:
            ws[f'A{row}'] = f"  - {ptype}"
            row += 1

        # Gaps and recommendations
        if self.gaps:
            row += 1
            ws[f'A{row}'] = "GAPS IDENTIFIED"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for gap in self.gaps.get("gaps", [])[:5]:  # Top 5 gaps
                ws[f'A{row}'] = f"  [{gap.get('severity', 'unknown').upper()}] {gap.get('description', '')}"
                row += 1

            row += 1
            ws[f'A{row}'] = "RECOMMENDATIONS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for rec in self.gaps.get("recommendations", [])[:5]:  # Top 5 recommendations
                ws[f'A{row}'] = f"  - {rec}"
                row += 1

        # Mental model insights
        if self.mental_model:
            row += 2
            ws[f'A{row}'] = "STATE ARCHITECTURE INSIGHTS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            arch = self.mental_model.get("program_architecture", "unknown")
            ws[f'A{row}'] = f"Architecture: {arch}"
            row += 1

            if self.mental_model.get("key_agencies"):
                ws[f'A{row}'] = "Key Agencies:"
                row += 1
                for agency in self.mental_model.get("key_agencies", [])[:5]:
                    ws[f'A{row}'] = f"  - {agency}"
                    row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _calculate_stats(self) -> DatabaseStats:
        """Calculate database statistics"""
        total = len(self.programs)

        def count_status(status: str) -> int:
            return sum(1 for p in self.programs
                      if p.get("status_tag", p.get("Status_Tag", "")) == status)

        active = count_status("ACTIVE")
        federal = count_status("FEDERAL")
        duplicate = count_status("DUPLICATE")
        expired = count_status("EXPIRED")
        non_incentive = count_status("NON-INCENTIVE")
        missing_link = count_status("MISSING-LINK")
        hallucination = count_status("HALLUCINATION")

        # Get coverage from gaps if available
        coverage_score = 0.0
        if self.gaps:
            coverage_score = self.gaps.get("coverage_score", 0.0)

        # Collect populations covered
        populations = set()
        for prog in self.programs:
            pops = prog.get("target_populations", [])
            if isinstance(pops, str):
                pops = [pops]
            populations.update(pops)

        # Collect program types
        program_types = set()
        for prog in self.programs:
            ptype = prog.get("benefit_type", prog.get("program_type", ""))
            if ptype:
                program_types.add(ptype)

        return DatabaseStats(
            total_programs=total,
            active_programs=active + federal,  # Both are usable
            federal_programs=federal,
            duplicate_programs=duplicate,
            expired_programs=expired,
            non_incentive_programs=non_incentive,
            missing_link_programs=missing_link,
            hallucination_programs=hallucination,
            coverage_score=coverage_score,
            populations_covered=sorted(list(populations)),
            program_types_found=sorted(list(program_types))
        )

    def _format_populations(self, populations: Any) -> str:
        """Format populations list as string"""
        if isinstance(populations, list):
            return ", ".join(populations)
        return str(populations) if populations else ""

    def _get_cleanup_guidance(self, status: str, program: Dict[str, Any]) -> tuple:
        """Get issue description and suggested action for cleanup items"""
        guidance = {
            "DUPLICATE": (
                "Duplicate of existing program",
                "Merge with original or remove"
            ),
            "NON-INCENTIVE": (
                "No direct employer benefit",
                "Verify classification or remove from incentive database"
            ),
            "MISSING-LINK": (
                "Cannot verify - no source URL",
                "Find official source URL or mark as unverifiable"
            ),
            "EXPIRED": (
                "Program no longer active",
                "Verify expiration or check for reauthorization"
            ),
            "HALLUCINATION": (
                "Likely fabricated program",
                "Remove from database unless can be verified"
            ),
            "REVIEW": (
                "Needs manual review",
                "Verify program details and update classification"
            )
        }

        issue, action = guidance.get(status, ("Unknown issue", "Review manually"))

        # Add reasoning if available
        reasoning = program.get("classification_reasoning", "")
        if reasoning:
            issue = f"{issue}: {reasoning}"

        return issue, action

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """Write a pandas DataFrame to a worksheet"""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _style_header(self, ws):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def _apply_status_colors(self, ws, status_col: int):
        """Apply color coding to status column"""
        status_colors = {
            "ACTIVE": "92D050",      # Green
            "FEDERAL": "00B0F0",     # Blue
            "DUPLICATE": "FFC000",   # Orange
            "EXPIRED": "808080",     # Gray
            "NON-INCENTIVE": "FFFF00",  # Yellow
            "MISSING-LINK": "FF6600",   # Dark orange
            "HALLUCINATION": "FF0000",  # Red
            "REVIEW": "9999FF"       # Purple
        }

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_col)
            status = cell.value
            if status in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )


def generate_summary(programs: List[Dict[str, Any]], gaps: Optional[Dict[str, Any]] = None) -> str:
    """
    Generate a text summary of the database.

    Returns formatted string suitable for printing or saving.
    """
    total = len(programs)

    def count_status(status: str) -> int:
        return sum(1 for p in programs
                  if p.get("status_tag", p.get("Status_Tag", "")) == status)

    active = count_status("ACTIVE")
    federal = count_status("FEDERAL")
    duplicate = count_status("DUPLICATE")
    expired = count_status("EXPIRED")
    non_incentive = count_status("NON-INCENTIVE")
    missing_link = count_status("MISSING-LINK")
    hallucination = count_status("HALLUCINATION")

    summary = f"""
{'='*60}
DATABASE SUMMARY
{'='*60}

Total Programs: {total}

BY STATUS:
  ACTIVE (usable):      {active}
  FEDERAL:              {federal}
  DUPLICATE:            {duplicate}
  EXPIRED:              {expired}
  NON-INCENTIVE:        {non_incentive}
  MISSING-LINK:         {missing_link}
  HALLUCINATION:        {hallucination}

USABLE PROGRAMS: {active + federal}
"""

    if gaps:
        coverage = gaps.get("coverage_score", 0)
        summary += f"""
COVERAGE SCORE: {coverage:.1f}%

GAPS IDENTIFIED: {len(gaps.get('gaps', []))}
"""
        for gap in gaps.get("gaps", [])[:3]:
            summary += f"  - [{gap.get('severity', 'unknown').upper()}] {gap.get('description', '')}\n"

        summary += f"""
RECOMMENDATIONS:
"""
        for rec in gaps.get("recommendations", [])[:3]:
            summary += f"  - {rec}\n"

    summary += f"\n{'='*60}\n"

    return summary
